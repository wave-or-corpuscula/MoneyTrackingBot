from datetime import datetime


from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment

from tgbot.utils import Database
from tgbot.models import Spending


class ExcelManager:

    reports_folder = "./tgbot/data/"

    headers = ["Категория", "Сумма", "Дата", "Описание"]

    start_row = 2
    start_col = 2

    middle_border = Border(
        left=Side(style="medium"), 
        bottom=Side(style="thin"), 
        right=Side(style="medium"))
    
    bottom_border = Border(
        left=Side(style="medium"), 
        bottom=Side(style="medium"), 
        right=Side(style="medium"))
    
    full_border = Border(
        left=Side(style="medium"), 
        bottom=Side(style="medium"), 
        right=Side(style="medium"), 
        top=Side(style="medium"))
    

    def create_excel_report(self, user_id: int, full_name: str, db: Database):
        # wb = load_workbook(f"{self.reports_folder}template.xlsx")
        wb = Workbook()
        wb.remove(wb["Sheet"])

        spendings = db.get_user_spendings_by_month(user_id)

        for spending in spendings.items():
            self.fill_month_spending_worksheet(wb, spending[0], spending[1])

        filename = f"{self.reports_folder}/{full_name} отчет трат - {datetime.now().date()}.xlsx"
        try:
            wb.save(filename)
        except OSError:
            filename = f"{self.reports_folder}/Отчет трат - {datetime.now().date()}.xlsx"
            wb.save(filename)
        return filename
    
    def fill_month_spending_worksheet(self, wb: Workbook, ws_name: str, spendings: list[Spending]):
        wb.create_sheet(ws_name)
        ws = wb[ws_name]

        for i, header in enumerate(self.headers):
            cell = ws.cell(row=self.start_row, column=self.start_col + i)
            cell.value = header
            cell.border = self.full_border
            cell.alignment = Alignment(horizontal="center")

        border = self.middle_border
        for i, spending in enumerate(spendings):
            spending_fields = [spending.spending_type_id.type_name, spending.spending, spending.spending_date, spending.description]
            for j, value in enumerate(spending_fields):
                if i == len(spendings) - 1:
                    border = self.bottom_border
                cell = ws.cell(row=self.start_row + i + 1, column=self.start_col + j)
                cell.value = value
                cell.border = border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    print("Excepted")
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
