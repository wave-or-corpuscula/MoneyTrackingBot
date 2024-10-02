from datetime import datetime

from aiogram import types

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from tgbot.utils import Database


class Multitool:

    @staticmethod
    def feature_in_process_text(data : types.Message | types.CallbackQuery):
        if isinstance(data, types.CallbackQuery):
            html_text = data.message.html_text
        else:
            html_text = data.html_text
        text = [
            "<u>Функция в разработке</u>\n",
            html_text,
        ]
        return "\n".join(text)
    

    @staticmethod
    def create_excel_report(user_id: int, full_name: str):
        reports_folder = "./tgbot/data/"

        wb = load_workbook(f"{reports_folder}template.xlsx")
        ws = wb.active

        middle_border = Border(left=Side(style="medium"), bottom=Side(style="thin"), right=Side(style="medium"))
        bottom_border = Border(left=Side(style="medium"), bottom=Side(style="medium"), right=Side(style="medium"))

        start_row = 3
        start_col = 2

        spendings = Database.get_user_spendings(user_id)

        border = middle_border
        for i, spending in enumerate(spendings):
            if i == len(spendings) - 1:
                border = bottom_border
            ws.cell(row=start_row + i, column=start_col, value=spending.spending_type_id.type_name).border = border
            ws.cell(row=start_row + i, column=start_col + 1, value=spending.spending).border = border
            ws.cell(row=start_row + i, column=start_col + 2, value=spending.spending_date).border = border
            ws.cell(row=start_row + i, column=start_col + 3, value=spending.description).border = border

        filename = f"{reports_folder}reports/{full_name} отчет трат - {datetime.now().date()}.xlsx"
        try:
            wb.save(filename)
        except OSError:
            filename = f"{reports_folder}reports/Отчет трат - {datetime.now().date()}.xlsx"
            wb.save(filename)
        return filename
